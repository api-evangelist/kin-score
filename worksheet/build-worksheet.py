#!/usr/bin/env python3
"""
Build the Kin Score Worksheet — a self-service .xlsx a provider uses to plan and
track work against the published rubric.

Source of truth: ../rubric/scoring-<version>.yml (the published snapshot) and the
arithmetic implemented in api-search/signals/score.rb. This script reproduces that
arithmetic in spreadsheet formulas; it does not invent checks, points or weights.

Usage:
    python3 build-worksheet.py [--rubric ../rubric/scoring-0.9.1.yml] [--out <path>]

Requires: openpyxl, PyYAML  (pip install openpyxl pyyaml)

WHEN THE RUBRIC BUMPS: re-run this. The only hand-maintained tables below are
EVAL_MODE (which checks the scorer evaluates against parsed specs vs. provider
frontmatter) and the gotcha copy in the Read Me — both need a look against
score.rb when checks are added.
"""

import argparse
import os
import re
import sys

try:
    import yaml
except ImportError:
    sys.exit("PyYAML required: pip install pyyaml")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")


# --------------------------------------------------------------------------
# Brand
# --------------------------------------------------------------------------
BLUE = "3098D8"          # API Evangelist brand primary (logo blue)
BLUE_DARK = "1C5F90"
BLUE_TINT = "E3F1FB"
GREY = "F4F6F8"
GREY_TEXT = "5A6570"
RULE = "D9DEE3"

H1 = Font(name="Helvetica Neue", size=16, bold=True, color=BLUE_DARK)
H2 = Font(name="Helvetica Neue", size=12, bold=True, color=BLUE_DARK)
HDR = Font(name="Helvetica Neue", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Helvetica Neue", size=10)
BODY_B = Font(name="Helvetica Neue", size=10, bold=True)
SMALL = Font(name="Helvetica Neue", size=9, color=GREY_TEXT)
MONO = Font(name="Menlo", size=9, color=GREY_TEXT)

FILL_HDR = PatternFill("solid", fgColor=BLUE)
FILL_TINT = PatternFill("solid", fgColor=BLUE_TINT)
FILL_GREY = PatternFill("solid", fgColor=GREY)
FILL_INPUT = PatternFill("solid", fgColor="FFF8E1")   # anything the user types

WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
CTR = Alignment(horizontal="center", vertical="top")
THIN = Border(bottom=Side(style="thin", color=RULE))


# --------------------------------------------------------------------------
# How score.rb classifies each check. :spec checks are read out of parsed
# OpenAPI/AsyncAPI documents and are N/A (out of the denominator) when the
# provider publishes none of that kind. :graded checks return a fraction and —
# note — stay IN the denominator whether or not a spec exists. :provider checks
# read the provider record and always count.
# --------------------------------------------------------------------------
SPEC_MODE = {
    "openapi_3_1", "openapi_3_0", "info_complete", "servers_defined",
    "operations_summary_coverage", "operations_description_coverage",
    "operations_tagged", "operations_operation_ids", "response_coverage",
    "error_responses_documented", "components_reuse", "security_schemes_defined",
    "security_applied", "oauth_flows_current", "credentials_not_in_query",
    "oauth_scopes_enumerated", "webhooks_or_callbacks",
    "asyncapi_3_x", "channels_defined", "messages_with_schema",
    "server_bindings", "security_schemes",
}
GRADED_MODE = {"servers_resolvable", "examples_present"}

# Provenance-graded checks: the award is multiplied by who authored the artifact.
# provenance.applies_to (per-check) plus provenance.applies_to_artifact (whole
# openapi block, from 0.8).
PROV_PER_CHECK = {
    "mcp_server_present": "mcp",
    "agent_skill_present": "skills",
    "contract_present": "openapi",
    "conformance_declared": "conformance",
}
PROV_ARTIFACT_BLOCK = {"openapi": "openapi"}

# Agent Readiness dimensions whose award is provenance-graded.
PROV_AGENT = {
    "mcp_server": "mcp",
    "agent_skills": "skills",
    "event_surface_described": "asyncapi",
    "agentic_access": "agentic_access",
}

FACET_ORDER = [
    "discoverability", "contract_quality", "governance",
    "operational_transparency", "developer_ergonomics",
    "commercial_clarity", "regulatory",
]

ARTIFACT_ORDER = [
    "apis_yml", "openapi", "asyncapi", "graphql", "fhir", "json_schema",
    "json_ld", "spectral", "operational", "ergonomics", "commercial", "regulatory",
]


def squash(text):
    return re.sub(r"\s+", " ", str(text or "")).strip()


def gate_for(artifact, check_id):
    """Which 'do you publish this?' toggle switches the check on."""
    if artifact == "graphql":
        return "graphql"
    if artifact == "fhir":
        return "fhir"
    if check_id in SPEC_MODE:
        return "asyncapi" if artifact == "asyncapi" else "openapi"
    return "always"


def load_rubric(path):
    with open(path) as fh:
        return yaml.safe_load(fh)


def build(rubric, out_path, version_note):
    facets = rubric["facets"]
    reg_weight = float(facets["regulatory"]["weight"])

    checks = []
    for art_name in ARTIFACT_ORDER:
        art = rubric["artifacts"][art_name]
        for c in art.get("checks") or []:
            regime = c.get("regime")
            if regime is None:
                regimes = ""
            elif isinstance(regime, list):
                regimes = ",".join(str(r) for r in regime)
            else:
                regimes = str(regime)
            prov = PROV_PER_CHECK.get(c["id"]) or PROV_ARTIFACT_BLOCK.get(art_name)
            checks.append({
                "artifact": art_name,
                "artifact_label": art.get("label", art_name),
                "id": c["id"],
                "label": c["label"],
                "facet": c["facet"],
                "points": int(c["points"]),
                "rule": squash(c.get("rule")),
                "why": squash(c.get("description")),
                "gate": gate_for(art_name, c["id"]),
                "regimes": regimes,
                "prov": prov or "",
                "mode": ("spec" if c["id"] in SPEC_MODE
                         else "graded" if c["id"] in GRADED_MODE else "provider"),
            })

    wb = Workbook()

    readme = wb.active
    readme.title = "Read Me"
    setup = wb.create_sheet("Setup")
    board = wb.create_sheet("Scorecard")
    sheet = wb.create_sheet("Checklist")
    agent = wb.create_sheet("Agent Readiness")
    prio = wb.create_sheet("Priorities")
    bands_ws = wb.create_sheet("Bands")
    regimes_ws = wb.create_sheet("Regulated Industries")
    prov_ws = wb.create_sheet("Provenance")

    n = len(checks)
    first, last = 2, n + 1

    # =====================================================================
    # SETUP
    # =====================================================================
    setup.sheet_view.showGridLines = False
    setup["A1"] = "Kin Score Worksheet — Setup"
    setup["A1"].font = H1
    setup["A2"] = ("Fill in the yellow cells. They switch whole blocks of checks on or off, "
                   "exactly the way the scorer does.")
    setup["A2"].font = SMALL

    rows = [
        ("A4", "Your API program", None),
        ("A5", "Provider name", ""),
        ("A6", "APIs.io provider slug", ""),
        ("A7", "Assessment date", ""),
        ("A8", "Assessed by", ""),
    ]
    for cell, label, val in rows:
        setup[cell] = label
        setup[cell].font = H2 if val is None else BODY
        if val is not None:
            c = setup[cell.replace("A", "B")]
            c.value = val
            c.fill = FILL_INPUT
            c.font = BODY

    setup["A10"] = "What you publish today"
    setup["A10"].font = H2
    setup["A11"] = ("A contract you do not publish is not counted against you — those checks leave the "
                    "denominator entirely. Answer honestly; overstating here inflates the whole sheet.")
    setup["A11"].font = SMALL

    toggles = [
        (12, "Do you publish an OpenAPI (or Swagger 2.0) contract?", "Yes",
         "Switches on the 17 spec-level OpenAPI checks."),
        (13, "Do you publish an AsyncAPI document?", "No",
         "Switches on the 5 spec-level AsyncAPI checks."),
        (14, "Do you publish a GraphQL SDL?", "No",
         "Switches on the 3 GraphQL checks. Not a GraphQL shop = not penalised."),
        (15, "Do you publish a FHIR CapabilityStatement?", "No",
         "Switches on the 2 FHIR checks. Not a FHIR server = not penalised."),
    ]
    for r, label, default, note in toggles:
        setup.cell(r, 1, label).font = BODY
        c = setup.cell(r, 2, default)
        c.fill = FILL_INPUT
        c.font = BODY_B
        c.alignment = CTR
        setup.cell(r, 3, note).font = SMALL

    setup["A17"] = "Regulated industry"
    setup["A17"].font = H2
    setup["A18"] = ("The regulatory facet is conditional. Pick the regime your tags match — see the "
                    "'Regulated Industries' tab — or leave it as none. Most-specific regime wins.")
    setup["A18"].font = SMALL
    setup["A19"] = "Regime"
    setup["A19"].font = BODY
    setup["B19"] = "none"
    setup["B19"].fill = FILL_INPUT
    setup["B19"].font = BODY_B
    setup["C19"] = (f"When set, the regulatory facet takes {reg_weight:.0%} of the composite and the "
                    f"other six scale to the remaining {1 - reg_weight:.0%}.")
    setup["C19"].font = SMALL

    # Gate lookup table (drives the Applies column on the Checklist).
    setup["E12"] = "gate"
    setup["F12"] = "applies"
    for c in ("E12", "F12"):
        setup[c].font = HDR
        setup[c].fill = FILL_HDR
    gate_rows = [("always", "Yes"), ("openapi", "=B12"), ("asyncapi", "=B13"),
                 ("graphql", "=B14"), ("fhir", "=B15")]
    for i, (g, v) in enumerate(gate_rows):
        setup.cell(13 + i, 5, g).font = MONO
        setup.cell(13 + i, 6, v).font = MONO

    # Provenance credit table.
    setup["E19"] = "authorship"
    setup["F19"] = "credit"
    for c in ("E19", "F19"):
        setup[c].font = HDR
        setup[c].fill = FILL_HDR
    credit = rubric["provenance"]["credit"]
    prov_rows = [("—", 1.0)] + [(k, float(v)) for k, v in credit.items()]
    for i, (k, v) in enumerate(prov_rows):
        setup.cell(20 + i, 5, k).font = MONO
        setup.cell(20 + i, 6, v).font = MONO

    # Regime id list, for the dropdown.
    setup["H12"] = "regime ids"
    setup["H12"].font = HDR
    setup["H12"].fill = FILL_HDR
    regime_ids = ["none"] + list(rubric["industry_regulatory"].keys())
    for i, rid in enumerate(regime_ids):
        setup.cell(13 + i, 8, rid).font = MONO

    for col, w in zip("ABCDEFGH", (46, 14, 62, 3, 22, 12, 3, 26)):
        setup.column_dimensions[col].width = w

    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    setup.add_data_validation(dv_yn)
    dv_yn.add("B12:B15")
    dv_regime = DataValidation(
        type="list", formula1=f"=Setup!$H$13:$H${12 + len(regime_ids)}", allow_blank=False)
    setup.add_data_validation(dv_regime)
    dv_regime.add("B19")

    # Named ranges keep the Checklist formulas readable.
    wb.defined_names.add(_defn("REGIME", "Setup!$B$19"))
    wb.defined_names.add(_defn("GATES", f"Setup!$E$13:$F${12 + len(gate_rows)}"))
    wb.defined_names.add(_defn("PROV_CREDIT", f"Setup!$E$20:$F${19 + len(prov_rows)}"))
    wb.defined_names.add(_defn("REGULATED", "Setup!$B$19"))

    # =====================================================================
    # CHECKLIST
    # =====================================================================
    headers = [
        ("Artifact", 16), ("Facet", 22), ("Check ID", 30), ("Check", 34),
        ("Points", 7), ("Applies", 9), ("Status", 10), ("Coverage", 10),
        ("Authorship", 13), ("Credit", 8), ("Earned", 8), ("Counts for", 10),
        ("Gap", 7), ("Worth on your score", 12), ("What the scorer looks for", 52),
        ("Why it matters", 76), ("Owner", 14), ("Target", 12), ("Notes", 34),
        ("_gate", 10), ("_regimes", 26), ("_rank", 10),
    ]
    _header_row(sheet, headers)
    sheet.freeze_panes = "E2"

    facet_labels = {k: v["label"] for k, v in facets.items()}

    for i, c in enumerate(checks):
        r = first + i
        sheet.cell(r, 1, c["artifact_label"]).font = BODY
        sheet.cell(r, 2, facet_labels[c["facet"]]).font = BODY
        sheet.cell(r, 3, c["id"]).font = MONO
        sheet.cell(r, 4, c["label"]).font = BODY_B
        sheet.cell(r, 5, c["points"]).font = BODY
        sheet.cell(r, 5).alignment = CTR

        sheet.cell(r, 6, (
            f'=IF(AND($U{r}<>"",REGIME="none"),"N/A",'
            f'IF(AND($U{r}<>"",$U{r}<>"any",ISERROR(SEARCH(REGIME,$U{r}))),"N/A",'
            f'IF(VLOOKUP($T{r},GATES,2,FALSE)="Yes","Yes","N/A")))'
        ))

        st = sheet.cell(r, 7, "No")
        st.fill = FILL_INPUT
        st.font = BODY
        st.alignment = CTR

        cov = sheet.cell(r, 8)
        cov.fill = FILL_INPUT
        cov.number_format = "0%"
        cov.alignment = CTR

        auth = sheet.cell(r, 9, c["prov"] and "first-party" or "—")
        auth.font = BODY
        auth.alignment = CTR
        if c["prov"]:
            auth.fill = FILL_INPUT

        sheet.cell(r, 10, f'=IFERROR(VLOOKUP($I{r},PROV_CREDIT,2,FALSE),1)').number_format = "0.00"
        sheet.cell(r, 11, (
            f'=IF($F{r}="N/A",0,$E{r}*IF(ISNUMBER($H{r}),MEDIAN(0,$H{r},1),'
            f'IF($G{r}="Yes",1,IF($G{r}="Partial",0.5,0)))*$J{r})'
        )).number_format = "0.0"
        sheet.cell(r, 12, f'=IF($F{r}="N/A",0,$E{r})').alignment = CTR
        sheet.cell(r, 13, f'=$L{r}-$K{r}').number_format = "0.0"
        sheet.cell(r, 14, (
            f'=IF($L{r}=0,0,ROUND($M{r}/SUMIF($B${first}:$B${last},$B{r},'
            f'$L${first}:$L${last})*100*'
            f"IFERROR(VLOOKUP($B{r},Scorecard!$A$6:$C$12,3,FALSE),0),2))"
        )).number_format = "0.00"

        sheet.cell(r, 15, c["rule"]).font = MONO
        sheet.cell(r, 15).alignment = WRAP
        sheet.cell(r, 16, c["why"]).font = BODY
        sheet.cell(r, 16).alignment = WRAP
        for col in (17, 18, 19):
            sheet.cell(r, col).fill = FILL_INPUT
        sheet.cell(r, 18).number_format = "yyyy-mm-dd"

        sheet.cell(r, 20, c["gate"]).font = MONO
        sheet.cell(r, 21, c["regimes"]).font = MONO
        sheet.cell(r, 22, f"=$N{r}-ROW()/100000")

        for col in range(1, 23):
            cell = sheet.cell(r, col)
            cell.border = THIN
            if cell.alignment == Alignment():
                cell.alignment = TOP

    dv_status = DataValidation(type="list", formula1='"Yes,Partial,No"', allow_blank=False)
    sheet.add_data_validation(dv_status)
    dv_status.add(f"G{first}:G{last}")
    dv_auth = DataValidation(
        type="list", formula1='"—,first-party,conformance,derived,unknown"', allow_blank=True)
    sheet.add_data_validation(dv_auth)
    dv_auth.add(f"I{first}:I{last}")

    sheet.conditional_formatting.add(
        f"F{first}:F{last}",
        CellIsRule(operator="equal", formula=['"N/A"'],
                   font=Font(name="Helvetica Neue", size=10, color="A0A8AE", italic=True)))
    sheet.conditional_formatting.add(
        f"G{first}:G{last}",
        CellIsRule(operator="equal", formula=['"Yes"'],
                   fill=PatternFill("solid", fgColor="D9F2E3")))

    sheet.column_dimensions["T"].hidden = True
    sheet.column_dimensions["U"].hidden = True
    sheet.column_dimensions["V"].hidden = True
    sheet.auto_filter.ref = f"A1:S{last}"

    # =====================================================================
    # SCORECARD
    # =====================================================================
    board.sheet_view.showGridLines = False
    board["A1"] = "Kin Score — Scorecard"
    board["A1"].font = H1
    board["A2"] = '=IF(Setup!B5="","(name your provider on the Setup tab)",Setup!B5)'
    board["A2"].font = H2
    board["A3"] = (f"Rubric {version_note}. Every number here is computed from the Checklist tab — "
                   "nothing on this sheet is typed.")
    board["A3"].font = SMALL

    _header_row(board, [
        ("Facet", 26), ("Weight", 10), ("Effective weight", 15),
        ("Points available", 15), ("Points earned", 14),
        ("Sub-score (0–100)", 16), ("Contribution", 13), ("", 4),
        ("What it measures", 78),
    ], row=5)

    for i, fname in enumerate(FACET_ORDER):
        r = 6 + i
        cfg = facets[fname]
        board.cell(r, 1, cfg["label"]).font = BODY_B
        board.cell(r, 2, float(cfg["weight"])).number_format = "0%"
        if fname == "regulatory":
            eff = f'=IF(REGIME="none",0,{reg_weight})'
        else:
            eff = f'=IF(REGIME="none",$B{r},$B{r}*{1 - reg_weight})'
        board.cell(r, 3, eff).number_format = "0.0%"
        board.cell(r, 4, f"=SUMIF(Checklist!$B${first}:$B${last},$A{r},"
                         f"Checklist!$L${first}:$L${last})")
        board.cell(r, 5, f"=SUMIF(Checklist!$B${first}:$B${last},$A{r},"
                         f"Checklist!$K${first}:$K${last})").number_format = "0.0"
        board.cell(r, 6, f'=IF($D{r}=0,0,ROUND($E{r}/$D{r}*100,1))').number_format = "0.0"
        board.cell(r, 7, f"=ROUND($C{r}*$F{r},2)").number_format = "0.00"
        board.cell(r, 9, squash(cfg["description"])).alignment = WRAP
        board.cell(r, 9).font = SMALL
        for col in range(1, 10):
            board.cell(r, col).border = THIN

    total = 13
    board.cell(total, 1, "COMPOSITE — your Kin Score").font = H2
    board.cell(total, 6, "=ROUND(SUM($G$6:$G$12),1)").font = Font(
        name="Helvetica Neue", size=20, bold=True, color=BLUE_DARK)
    board.cell(total, 6).alignment = CTR
    board.cell(total, 6).number_format = "0.0"
    for col in range(1, 8):
        board.cell(total, col).fill = FILL_TINT

    bands = rubric["bands"]
    band_formula = _band_formula(bands, "$F$13")
    board.cell(14, 1, "Band").font = BODY_B
    board.cell(14, 6, band_formula).font = H2
    board.cell(14, 6).alignment = CTR
    board.cell(14, 9, "=VLOOKUP($F$14,Bands!$A$4:$D$9,4,FALSE)").font = SMALL
    board.cell(14, 9).alignment = WRAP

    board.cell(16, 1, "Where the next points are").font = H2
    board.cell(17, 1, "Checks passing").font = BODY
    board.cell(17, 6, f'=COUNTIFS(Checklist!$F${first}:$F${last},"Yes",'
                      f'Checklist!$G${first}:$G${last},"Yes")').alignment = CTR
    board.cell(18, 1, "Checks that apply to you").font = BODY
    board.cell(18, 6, f'=COUNTIF(Checklist!$F${first}:$F${last},"Yes")').alignment = CTR
    board.cell(19, 1, "Checks not applicable").font = BODY
    board.cell(19, 6, f'=COUNTIF(Checklist!$F${first}:$F${last},"N/A")').alignment = CTR
    board.cell(20, 1, "Composite points still on the table").font = BODY
    board.cell(20, 6, f"=ROUND(SUM(Checklist!$N${first}:$N${last}),1)").alignment = CTR
    board.cell(20, 6).font = BODY_B
    board.cell(21, 1, "Biggest single win").font = BODY
    board.cell(21, 6, "=Priorities!$D$4").alignment = CTR
    board.cell(21, 6).number_format = "0.00"
    board.cell(21, 9, '=IF(Priorities!$B$4="","nothing left that applies to you",'
                      'Priorities!$B$4&"  ("&Priorities!$C$4&")")').font = SMALL

    board.cell(23, 1, "Agent Readiness (scored separately, never blended in)").font = H2
    board.cell(24, 1, "Score").font = BODY
    board.cell(24, 6, "='Agent Readiness'!$C$2").alignment = CTR
    board.cell(24, 6).font = Font(name="Helvetica Neue", size=14, bold=True, color=BLUE_DARK)
    board.cell(25, 1, "Band").font = BODY
    board.cell(25, 6, "='Agent Readiness'!$C$3").alignment = CTR
    board.cell(25, 6).font = BODY_B
    board.cell(25, 9, ("A composite in the Strong band and a human-only Agent Readiness is a normal, "
                       "and increasingly expensive, combination.")).font = SMALL

    for col, w in zip("ABCDEFGHI", (34, 10, 15, 15, 14, 16, 13, 4, 78)):
        board.column_dimensions[col].width = w

    # =====================================================================
    # AGENT READINESS
    # =====================================================================
    agent.sheet_view.showGridLines = False
    agent["A1"] = "Agent Readiness"
    agent["A1"].font = H1
    ar = rubric["agent_readiness"]
    dims = ar["dimensions"]
    ar_first, ar_last = 7, 6 + len(dims)

    agent["A2"] = "Score (0–100)"
    agent["A2"].font = BODY_B
    agent["C2"] = f"=IF(SUM($C${ar_first}:$C${ar_last})=0,0,ROUND(SUM($J${ar_first}:$J${ar_last})/SUM($C${ar_first}:$C${ar_last})*100,1))"
    agent["C2"].font = Font(name="Helvetica Neue", size=16, bold=True, color=BLUE_DARK)
    agent["A3"] = "Band"
    agent["A3"].font = BODY_B

    idem_row = ar_first + [d["id"] for d in dims].index("idempotency")
    err_row = ar_first + [d["id"] for d in dims].index("error_semantics")
    ar_bands = ar["bands"]
    base_band = _band_formula(ar_bands, "$C$2", label_key="label")
    agent["C3"] = (
        f'=IF(AND({base_band[1:]}="{ar_bands[0]["label"]}",'
        f'OR($H{idem_row}=0,$H{err_row}=0)),"{ar_bands[1]["label"]}",{base_band[1:]})'
    )
    agent["C3"].font = H2
    agent["A4"] = squash(ar["description"])
    agent["A4"].font = SMALL
    agent["A5"] = (f'Band gate: {ar_bands[0]["label"]} is withheld unless idempotency AND stable error '
                   f'semantics both score. Every dimension is always in the denominator — unlike the '
                   f'composite, a missing contract here is a real deficiency, not an N/A.')
    agent["A5"].font = SMALL

    _header_row(agent, [
        ("Dimension", 30), ("ID", 24), ("Points", 7), ("Grades available", 34),
        ("Your grade", 14), ("Documented credit", 9), ("Authorship", 13),
        ("Credit", 8), ("Prov. credit", 9), ("Earned", 8), ("Gap", 7),
        ("What the scorer looks for", 60), ("Why it matters", 76),
        ("Owner", 14), ("Target", 12), ("Notes", 34),
    ], row=6)

    for i, d in enumerate(dims):
        r = ar_first + i
        graded = bool(d.get("graded"))
        cr = d.get("credit") or {}
        opts = ["no"] + [k for k in cr] if graded else ["no", "yes"]
        doc_credit = float(cr.get("documented", 0.5)) if graded else 0

        agent.cell(r, 1, d["label"]).font = BODY_B
        agent.cell(r, 2, d["id"]).font = MONO
        agent.cell(r, 3, int(d["points"])).alignment = CTR
        agent.cell(r, 4, ", ".join(opts)).font = SMALL
        g = agent.cell(r, 5, "no")
        g.fill = FILL_INPUT
        g.alignment = CTR
        agent.cell(r, 6, doc_credit).font = SMALL
        agent.cell(r, 6).alignment = CTR

        pclass = PROV_AGENT.get(d["id"], "")
        a = agent.cell(r, 7, "first-party" if pclass else "—")
        a.alignment = CTR
        if pclass:
            a.fill = FILL_INPUT

        agent.cell(r, 8, (
            f'=IF($E{r}="no",0,IF(OR($E{r}="yes",$E{r}="verified",$E{r}="conformant"),1,'
            f'IF($E{r}="near-conformant",0.6,IF($E{r}="partial",0.5,'
            f'IF($E{r}="documented",$F{r},IF($E{r}="flavored",0.25,0))))))'
        )).number_format = "0.00"
        agent.cell(r, 9, f'=IFERROR(VLOOKUP($G{r},PROV_CREDIT,2,FALSE),1)').number_format = "0.00"
        agent.cell(r, 10, f"=$C{r}*$H{r}*$I{r}").number_format = "0.0"
        agent.cell(r, 11, f"=$C{r}-$J{r}").number_format = "0.0"
        agent.cell(r, 12, squash(d.get("signal"))).font = MONO
        agent.cell(r, 12).alignment = WRAP
        agent.cell(r, 13, squash(d.get("description"))).font = BODY
        agent.cell(r, 13).alignment = WRAP
        for col in (14, 15, 16):
            agent.cell(r, col).fill = FILL_INPUT
        agent.cell(r, 15).number_format = "yyyy-mm-dd"

        dv = DataValidation(type="list", formula1='"' + ",".join(opts) + '"', allow_blank=False)
        agent.add_data_validation(dv)
        dv.add(f"E{r}")

        for col in range(1, 17):
            agent.cell(r, col).border = THIN
            if agent.cell(r, col).alignment == Alignment():
                agent.cell(r, col).alignment = TOP

    tot = ar_last + 1
    agent.cell(tot, 1, "TOTAL").font = BODY_B
    agent.cell(tot, 3, f"=SUM($C${ar_first}:$C${ar_last})").font = BODY_B
    agent.cell(tot, 10, f"=ROUND(SUM($J${ar_first}:$J${ar_last}),1)").font = BODY_B
    agent.cell(tot, 11, f"=ROUND(SUM($K${ar_first}:$K${ar_last}),1)").font = BODY_B
    for col in range(1, 12):
        agent.cell(tot, col).fill = FILL_TINT

    for col, w in zip("ABCDEFGHIJKLMNOP",
                      (30, 24, 7, 34, 14, 9, 13, 8, 9, 8, 7, 60, 76, 14, 12, 34)):
        agent.column_dimensions[col].width = w
    agent.column_dimensions["F"].hidden = True
    agent.freeze_panes = "C7"

    # =====================================================================
    # PRIORITIES
    # =====================================================================
    prio.sheet_view.showGridLines = False
    prio["A1"] = "Do these next"
    prio["A1"].font = H1
    top_n = 30
    _header_row(prio, [
        ("#", 5), ("Check", 34), ("Facet", 24), ("Worth on your score", 12),
        ("Points", 7), ("Artifact", 16), ("What the scorer looks for", 60),
        ("Owner", 14), ("Target", 12), ("_key", 10),
    ], row=3)
    prio["A2"] = ("Ranked by how much composite score each unfinished check is worth to YOU right now — "
                  "its points, divided by the denominator of its facet, times that facet's weight. "
                  "Recalculates as you tick things off on the Checklist. Small facets move fast: a "
                  "6-point governance check can outrank a 20-point contract check.")
    prio["A2"].font = SMALL
    prio["A2"].alignment = WRAP

    for i in range(top_n):
        r = 4 + i
        key = f'=LARGE(Checklist!$V${first}:$V${last},{i + 1})'
        prio.cell(r, 10, key)
        m = f'MATCH($J{r},Checklist!$V${first}:$V${last},0)'
        prio.cell(r, 1, i + 1).alignment = CTR
        prio.cell(r, 2, f'=IF($D{r}<=0,"",INDEX(Checklist!$D${first}:$D${last},{m}))').font = BODY_B
        prio.cell(r, 3, f'=IF($D{r}<=0,"",INDEX(Checklist!$B${first}:$B${last},{m}))')
        prio.cell(r, 4, f'=ROUND(MAX(0,INDEX(Checklist!$N${first}:$N${last},{m})),2)')
        prio.cell(r, 4).number_format = "0.00"
        prio.cell(r, 4).font = BODY_B
        prio.cell(r, 5, f'=IF($D{r}<=0,"",INDEX(Checklist!$E${first}:$E${last},{m}))').alignment = CTR
        prio.cell(r, 6, f'=IF($D{r}<=0,"",INDEX(Checklist!$A${first}:$A${last},{m}))')
        prio.cell(r, 7, f'=IF($D{r}<=0,"",INDEX(Checklist!$O${first}:$O${last},{m}))').font = MONO
        prio.cell(r, 7).alignment = WRAP
        for col in (8, 9):
            prio.cell(r, col).fill = FILL_INPUT
        prio.cell(r, 9).number_format = "yyyy-mm-dd"
        for col in range(1, 10):
            prio.cell(r, col).border = THIN
            if prio.cell(r, col).alignment == Alignment():
                prio.cell(r, col).alignment = TOP

    prio.column_dimensions["J"].hidden = True
    for col, w in zip("ABCDEFGHI", (5, 34, 24, 12, 7, 16, 60, 14, 12)):
        prio.column_dimensions[col].width = w

    # =====================================================================
    # BANDS
    # =====================================================================
    bands_ws.sheet_view.showGridLines = False
    bands_ws["A1"] = "Bands"
    bands_ws["A1"].font = H1
    bands_ws["A2"] = "Composite — the headline Kin Score"
    bands_ws["A2"].font = H2
    _header_row(bands_ws, [("Band", 16), ("Range", 16), ("Share of catalog", 15),
                           ("What it means", 96)], row=3)
    for i, b in enumerate(bands):
        r = 4 + i
        bands_ws.cell(r, 1, b["label"]).font = BODY_B
        bands_ws.cell(r, 2, b.get("range", "")).alignment = CTR
        bands_ws.cell(r, 3, b.get("share", "")).alignment = CTR
        bands_ws.cell(r, 4, squash(b["description"])).alignment = WRAP
        for col in range(1, 5):
            bands_ws.cell(r, col).border = THIN

    r0 = 4 + len(bands) + 1
    bands_ws.cell(r0, 1, "Agent Readiness — scored and displayed separately").font = H2
    _header_row(bands_ws, [("Band", 16), ("Range", 16), ("Share of catalog", 15),
                           ("What it means", 96)], row=r0 + 1)
    for i, b in enumerate(ar_bands):
        r = r0 + 2 + i
        bands_ws.cell(r, 1, b["label"]).font = BODY_B
        bands_ws.cell(r, 2, b.get("range", f'{b.get("min","")}+')).alignment = CTR
        bands_ws.cell(r, 3, b.get("share", "")).alignment = CTR
        bands_ws.cell(r, 4, squash(b.get("description"))).alignment = WRAP
        for col in range(1, 5):
            bands_ws.cell(r, col).border = THIN
    for col, w in zip("ABCD", (16, 16, 15, 96)):
        bands_ws.column_dimensions[col].width = w

    # =====================================================================
    # REGULATED INDUSTRIES
    # =====================================================================
    regimes_ws.sheet_view.showGridLines = False
    regimes_ws["A1"] = "Regulated industries"
    regimes_ws["A1"].font = H1
    regimes_ws["A2"] = ("A provider is regulated when its tags match a regime. Every matching regime is "
                        "scored and the one with the most tag hits wins, ties broken by specificity. "
                        "Pick the winner on the Setup tab — it switches on that regime's own checks and "
                        "folds the regulatory facet into your composite.")
    regimes_ws["A2"].font = SMALL
    regimes_ws["A2"].alignment = WRAP
    _header_row(regimes_ws, [("Regime ID", 24), ("Label", 26), ("Specificity", 10),
                             ("Regimes it stands for", 52), ("Tags that match it", 62),
                             ("Standards it expects", 52), ("Basis", 76)], row=4)
    for i, (rid, cfg) in enumerate(rubric["industry_regulatory"].items()):
        r = 5 + i
        regimes_ws.cell(r, 1, rid).font = MONO
        regimes_ws.cell(r, 2, cfg["label"]).font = BODY_B
        regimes_ws.cell(r, 3, cfg.get("specificity", "")).alignment = CTR
        regimes_ws.cell(r, 4, ", ".join(cfg.get("regimes", []))).alignment = WRAP
        tags = [str(t) for t in cfg.get("tags", [])]
        weak = [str(t) for t in cfg.get("weak_tags", [])]
        tag_txt = ", ".join(tags)
        if weak:
            tag_txt += "   |   weak (only when nothing else matched): " + ", ".join(weak)
        regimes_ws.cell(r, 5, tag_txt).alignment = WRAP
        regimes_ws.cell(r, 6, ", ".join(cfg.get("standards", []))).alignment = WRAP
        regimes_ws.cell(r, 7, squash(cfg.get("basis"))).alignment = WRAP
        for col in range(1, 8):
            regimes_ws.cell(r, col).border = THIN
    for col, w in zip("ABCDEFG", (24, 26, 10, 52, 62, 52, 76)):
        regimes_ws.column_dimensions[col].width = w

    # =====================================================================
    # PROVENANCE
    # =====================================================================
    prov_ws.sheet_view.showGridLines = False
    prov_ws["A1"] = "Provenance — who published the artifact"
    prov_ws["A1"].font = H1
    prov_ws["A2"] = (
        "API Evangelist's enrichment pipeline authors artifacts on a provider's behalf — modeled "
        "OpenAPI, candidate MCP catalogs, generated agent skills, derived agentic-access contracts. "
        "From rubric 0.6 the score can see the difference, and grades an artifact by who wrote it. "
        "On the Checklist and Agent Readiness tabs, the Authorship column is only editable on the "
        "checks this applies to; everything else is fixed at full credit."
    )
    prov_ws["A2"].font = SMALL
    prov_ws["A2"].alignment = WRAP
    _header_row(prov_ws, [("Authorship", 18), ("Credit", 10), ("What it means", 96)], row=4)
    meanings = {
        "first-party": "The provider published it themselves. Full credit.",
        "conformance": "Written against a standard the provider conforms to, not by the provider. Most of the award.",
        "derived": "API Evangelist authored it on the provider's behalf. A quarter credit — it is a real "
                   "artifact an agent can fetch, but it is not evidence the provider designed for agents.",
        "unknown": "Nobody marked it. Credited in full TODAY because marker coverage across the catalog is "
                   "still ~2.6% and the alternative punishes providers for a gap in our metadata. This is "
                   "the loosest dial in the rubric and it is expected to tighten — publishing an artifact "
                   "with legible authorship is the cheapest way to protect points you already have.",
    }
    for i, (k, v) in enumerate(credit.items()):
        r = 5 + i
        prov_ws.cell(r, 1, k).font = BODY_B
        prov_ws.cell(r, 2, float(v)).number_format = "0.00"
        prov_ws.cell(r, 2).alignment = CTR
        prov_ws.cell(r, 3, meanings.get(k, "")).alignment = WRAP
        for col in range(1, 4):
            prov_ws.cell(r, col).border = THIN

    r0 = 5 + len(credit) + 1
    prov_ws.cell(r0, 1, "Which checks are graded this way").font = H2
    _header_row(prov_ws, [("Check / dimension", 34), ("Artifact class", 18),
                          ("Where", 20), ("Points at stake", 96)], row=r0 + 1)
    graded_rows = []
    for c in checks:
        if c["prov"]:
            graded_rows.append((c["label"], c["prov"], "Checklist", c["points"]))
    for d in dims:
        if d["id"] in PROV_AGENT:
            graded_rows.append((d["label"], PROV_AGENT[d["id"]], "Agent Readiness", d["points"]))
    for i, (label, cls, where, pts) in enumerate(graded_rows):
        r = r0 + 2 + i
        prov_ws.cell(r, 1, label).font = BODY
        prov_ws.cell(r, 2, cls).font = MONO
        prov_ws.cell(r, 3, where).font = SMALL
        prov_ws.cell(r, 4, f"{pts} points, scaled by authorship credit").font = SMALL
        for col in range(1, 5):
            prov_ws.cell(r, col).border = THIN
    for col, w in zip("ABCD", (34, 18, 20, 96)):
        prov_ws.column_dimensions[col].width = w

    # =====================================================================
    # READ ME
    # =====================================================================
    readme.sheet_view.showGridLines = False
    readme.column_dimensions["A"].width = 4
    readme.column_dimensions["B"].width = 118

    def line(r, text, style=BODY, height=None):
        c = readme.cell(r, 2, text)
        c.font = style
        c.alignment = WRAP
        if height:
            readme.row_dimensions[r].height = height

    line(2, "Kin Score Worksheet", H1)
    line(3, f"Rubric {version_note} — the API Evangelist rating system, licensed by APIs.io", SMALL)
    line(5, "What this is", H2)
    line(6,
         "A working copy of the Kin Score rubric you can fill in for your own API program. It holds "
         f"every one of the {n} checks in the composite score and all {len(dims)} Agent Readiness "
         "dimensions, with their real point values, and it reproduces the scorer's arithmetic in "
         "formulas — including the conditional regulatory facet, the N/A rules for contracts you "
         "don't publish, and the provenance multiplier. Tick a check off and every number moves.",
         BODY, 62)
    line(8, "How to use it", H2)
    line(9, "1.  Setup tab — name your program and answer the four 'what do you publish' questions. "
            "Pick your regulatory regime, or leave it as none.", BODY)
    line(10, "2.  Checklist tab — work down it and set Status to Yes / Partial / No. Yellow cells are "
             "yours; everything else calculates. Use the Coverage column when a check is true of some "
             "of your contracts but not all — the scorer averages spec checks across every document "
             "you publish, so 6 of 10 specs with good summaries is 60%, not a pass.", BODY, 48)
    line(11, "3.  Priorities tab — reads back the highest-value unfinished work, ranked by what each "
             "check is actually worth to your score today.", BODY)
    line(12, "4.  Scorecard tab — your composite, your band, and where the remaining points sit.", BODY)
    line(13, "5.  Agent Readiness tab — scored separately and never blended into the composite. "
             "Several dimensions are graded rather than pass/fail; the grades available are listed "
             "on each row.", BODY, 34)

    line(15, "The arithmetic", H2)
    line(16, "Each check awards points to one facet. A facet's sub-score is points earned ÷ points "
             "available for that facet, as a 0–100. The composite is the weighted sum of the facet "
             "sub-scores.", BODY, 34)
    line(17, "Points available is not a constant. A check that cannot apply to you leaves the "
             "denominator instead of scoring zero — that is why the facet weights matter less than "
             "the size of the facet you are working in. This is the single most useful thing in the "
             "sheet: a 5-point check in a small facet can be worth more composite score than a "
             "20-point check in a large one, and the 'Worth on your score' column tells you which.",
         BODY, 62)
    line(18, f"The regulatory facet is conditional. Unregulated, it does not exist for you and the "
             f"other six facets carry the whole score. Regulated, it takes {reg_weight:.0%} and the "
             f"other six scale to {1 - reg_weight:.0%}. Regime-specific checks only apply inside "
             f"their regime.", BODY, 34)
    line(19, "Provenance multiplies the award, it does not gate it. An artifact API Evangelist "
             "authored on your behalf still earns a quarter of its points — see the Provenance tab.",
         BODY, 34)

    line(21, "What this sheet cannot do", H2)
    line(22, "It cannot set your real score. The published Kin Score is computed by the scorer against "
             "artifacts it can actually fetch and parse; this is a planning model of the same rubric. "
             "Filling it in optimistically changes nothing but the spreadsheet.", BODY, 34)
    line(23, "Four ways providers lose points for work they genuinely did — worth checking before you "
             "conclude the score is wrong:", BODY, 20)
    line(24, "•  A contract that is only linked, never harvested. If your apis.yml points 'type: "
             "OpenAPI' at a remote URL and nothing lands on disk, all 22 spec-level checks — 86 points "
             "— score against nothing. Publish the spec at a stable path and point the index at it.",
         BODY, 48)
    line(25, "•  An A2A agent card at a filename the scorer doesn't look for. It reads "
             "a2a/<slug>-a2a.yml, on your provider slug, not your company name.", BODY, 34)
    line(26, "•  A self-hosted index that isn't. 'apis.json self-hosted' rewards YOUR host — an index "
             "served from github, apievangelist.com or apis.io fails it by design. The check exists to "
             "reward the difference between describing yourself and being described.", BODY, 48)
    line(27, "•  Unmarked authorship. An artifact with no provenance marker is credited in full today, "
             "but that is the loosest dial in the rubric and it is expected to tighten. Marking your "
             "own work as yours protects points you already hold.", BODY, 48)

    line(29, "Where the real thing lives", H2)
    line(30, "Rubric, changelog and every versioned snapshot:  github.com/api-evangelist/kin-score", BODY)
    line(31, "Your live score and band:  apis.io  ·  providers.apievangelist.com", BODY)
    line(32, "Questions, or a score you think is wrong:  kin@apievangelist.com", BODY)
    line(34, "Every input to this score is a public, machine-checkable signal you chose to publish or "
             "chose to leave out. No human quietly adds or removes points.", SMALL, 30)

    wb.active = 0
    wb.save(out_path)
    return n, len(dims)


def _defn(name, ref):
    from openpyxl.workbook.defined_name import DefinedName
    return DefinedName(name, attr_text=ref)


def _header_row(ws, headers, row=1):
    for i, (title, width) in enumerate(headers, start=1):
        c = ws.cell(row, i, title)
        c.font = HDR
        c.fill = FILL_HDR
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 28


def _band_formula(bands, ref, label_key="label"):
    """Nested IF from a bands list ordered high to low."""
    ordered = sorted(bands, key=lambda b: -float(b.get("min", 0)))
    out = f'"{ordered[-1][label_key]}"'
    for b in ordered[:-1][::-1]:
        out = f'IF({ref}>={b["min"]},"{b[label_key]}",{out})'
    return "=" + out


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser()
    ap.add_argument("--rubric", default=os.path.join(here, "..", "rubric", "scoring-0.9.1.yml"))
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    rubric = load_rubric(args.rubric)
    version = str(rubric["schema_version"])
    note = f"{version} ({rubric.get('last_updated')})"
    out = args.out or os.path.join(here, f"kin-score-worksheet-{version}.xlsx")

    n, d = build(rubric, out, note)
    print(f"wrote {out}")
    print(f"  rubric {note} — {n} composite checks, {d} agent-readiness dimensions")


if __name__ == "__main__":
    main()
